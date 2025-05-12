import os
import sys
import json
import pandas as pd
import numpy as np
import select
from pathlib import Path
from openpyxl.styles import PatternFill
from openpyxl.styles.fonts import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from .commom import *

# fill = PatternFill(patternType='solid', fgColor='d3d3d3')
fill = PatternFill(fgColor='EEEEEE',bgColor="EEEEEE", fill_type = "solid")
font = Font(color="FF0000", bold = True)
side = Side(style='thin', color='000000')
border = Border(top=side, bottom=side, left=side, right=side)
alignment = Alignment(horizontal="center", vertical="center", wrapText = True)

class ADD_QC :

    def __init__(self, args) :
        self.flowcellid = args.flowcellid
        self.directory = args.directory
        self.project_type = args.project_type
        self.outdir = args.outdir

    def get_wet_info(self, fc_id, prj_type):
        try :
            connection = pymysql.connect(host="192.168.9.100", user="gxd_pipeline", password="gw!2341234", database="gxd")
        except Exception as e:
            sys.exit({e})

        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            db_tbl = pd.read_sql(getattr(self, f"SelectQC_WET_{prj_type}")(fc_id), connection)
        return db_tbl

    def SelectQC_WET_WTS(self, fc_id):
        query = f'''
        select DISTINCT l.sample_id AS SAMPLE_ID, COALESCE(l.ref_prep_id, l.prep_id) AS prep_id,
        COALESCE(p1.grade,p2.grade) AS PREP_GRADE,
        pre.grade AS PRE_GRADE,
        post.grade AS POST_GRADE
        from gxd.tb_order_line l
        left join tb_expr_prep p1
        on l.ref_prep_id = p1.prep_id
        left join tb_expr_prep p2
        on l.prep_id = p2.prep_id
        left join tb_expr_pre_pcr pre
        on l.sample_id = pre.sample_id
        left join tb_expr_post_pcr post
        on l.sample_id = post.sample_id
        left join gc_qc_sample q
        on l.sample_id = q.sample_id
        WHERE l.sample_id regexp '^CR_' AND q.FC_ID = '{fc_id}'
        '''
        return query

    def SelectQC_WET_eWES(self, fc_id):
        query = f'''
        select DISTINCT l.sample_id AS SAMPLE_ID, COALESCE(l.ref_prep_id, l.prep_id) AS prep_id,
        -- sample prep
        COALESCE(p1.grade,p2.grade) AS PREP_GRADE,
        -- library prep
        pre.grade AS PRE_GRADE,
        post.grade AS POST_GRADE
        from tb_order_line l
        left join tb_expr_prep p1
        on l.ref_prep_id = p1.prep_id
        left join tb_expr_prep p2
        on l.prep_id = p2.prep_id
        left join tb_expr_pre_pcr pre
        on l.sample_id = pre.sample_id
        left join tb_expr_post_pcr post
        on l.sample_id = post.sample_id
        left join gc_qc_sample q
        on l.sample_id = q.sample_id
        WHERE l.sample_id regexp '^CD_' AND q.FC_ID = '{fc_id}'
        '''
        return query

    def check_qc(self, df, tbl_th, parentDir):

        count_i = df.shape[0]
        for i, item in df.iterrows() :
            FILE = '/'.join([parentDir, item['PRJ_TYPE'], item['seqDir'], item['SAMPLE_ID'], 'Summary', '.'.join([item['SAMPLE_ID'],'report','json'])])
            try :
                DATA = json.load(open(FILE, 'r'))

                if "df_qc" in locals():
                    df_qc = pd.concat([df_qc, pd.DataFrame(DATA['biData'], index=[item['SAMPLE_ID']])], axis=0)
                else :
                    df_qc = pd.DataFrame(DATA['biData'], index=[item['SAMPLE_ID']])

            except Exception as e:

                print('Sample ID: ' + item['SAMPLE_ID'] + ' analysis has not been completed.')
                print("Do you want to continue? (Y/N): ", end='', flush=True)
                rlist, _, _ = select.select([sys.stdin], [], [], 60)

                if rlist :
                    answer = sys.stdin.readline().strip().lower()
                    if answer in ['y', 'yes']:
                        print("Continuing the process...")
                        count_i += -1
                        if "df_qc" in locals():
                            df_qc = pd.concat([df_qc, pd.DataFrame(index=[item['SAMPLE_ID']])], axis=0)
                        else :
                            df_qc = pd.DataFrame(index=[item['SAMPLE_ID']])

                    elif answer in ['n', 'no']:
                        init("Process aborted by the user.")
                    else:
                        init("Invalid input. Process aborted.")

                else :
                    init("\nNo input received within 60 seconds. The process will now be aborted.")

        if count_i == 0: init('There are no samples completed for analysis.')

        tbl_th = tbl_th[tbl_th['PRJ_TYPE']==df.loc[0,'PRJ_TYPE']].reset_index(drop=True)

        comm_columns = list(set(tbl_th['INDEX']).intersection(df_qc.columns))
        tbl_th = tbl_th[tbl_th['INDEX'].isin(comm_columns)]
        df_qc = df_qc[comm_columns]
        if tbl_th.shape[0] != df_qc.shape[1] : return None

        out_columns = []
        for i, item in tbl_th.iterrows() :
            if item['HL'] == 'high':
                df_qc[item['TITLE'] + '_pass'] = [float(x) >= float(item['TH']) for x in df_qc[item['INDEX']]]
            elif item['HL'] == 'low':
                df_qc[item['TITLE'] + '_pass'] = [float(x) <= float(item['TH']) for x in df_qc[item['INDEX']]]
            out_columns.extend([item['TITLE'], item['TITLE'] + '_pass'])

        dict_col = dict(zip(tbl_th['INDEX'],tbl_th['TITLE']))
        df_qc.rename(columns=dict_col, inplace=True)
        df_qc = df_qc[out_columns].sort_index()

        return df_qc

    def __call__(self):

        thfile = os.path.join(os.path.dirname(__file__), 'thresholds.tsv')
        assert os.path.isfile(thfile)
        tbl_th = pd.read_csv(thfile, sep="\t", index_col=None)

        df_info = getinfo(self.flowcellid)
        if df_info.shape[0] == 0 : init("flowcell ID not registered.")

        df_info['PRJ_TYPE'] = df_info['PRJ_TYPE'].str.replace('EWES',"eWES")
        if self.project_type == "both" :
            df_info = df_info[ df_info['PRJ_TYPE'].isin(['eWES','WTS']) ]
        else :
            df_info = df_info[ df_info['PRJ_TYPE']==self.project_type]
        if df_info.shape[0] == 0 : init("No data for the relevant project type.")

        uniq_info = fcDir_table(df_info, self.directory)
        if uniq_info.shape[0] == 0: init("No analysed samples.")

        df_info = pd.merge(df_info, uniq_info, on=['sub_name','PRJ_TYPE'])

        for i, item in uniq_info.iterrows() :
            if item['PRJ_TYPE'] == 'eWES':
                PC = '_PCE_'
                NC = '_NCE_'
            elif item['PRJ_TYPE'] == 'WTS':
                PC = '_PCT_'
                NC = '_NCT_'
            else :
                print('PRJ_TYPE error:' + item['PRJ_TYPE'] )
                continue

            temp_info = df_info[ (df_info['sub_name']==item['sub_name']) & (df_info['PRJ_TYPE']==item['PRJ_TYPE']) ].reset_index(drop=True)
            temp_info = temp_info[['run_id','PRJ_TYPE','seqDir','ANAL_STATUS','SAMPLE_ID']]
            temp_qc = self.check_qc(temp_info, tbl_th, self.directory)
            temp_qc['CTRL'] = np.where(temp_qc.index.str.contains(PC, case=False, regex=False),'PC', np.where(temp_qc.index.str.contains(NC, case=False, regex=False),'NC',''))

            wet_info = self.get_wet_info(self.flowcellid, item['PRJ_TYPE'])
            temp_qc = pd.merge(wet_info, temp_qc, left_on='SAMPLE_ID', right_index=True, how='right')
            temp_qc = temp_qc.set_index('SAMPLE_ID')
            temp_qc.index.name = None
            temp_qc = temp_qc.assign(DRY='', TS='', GS='', LD='')

            if temp_qc.shape[1] > 0 :

                out_file = os.path.join(self.outdir, '.'.join([item['seqDir'], item['PRJ_TYPE'],'xlsx']))
                if not os.path.isfile(out_file) :
                    print (out_file + " is not exists. Create a file with create_worksheet.py.")
                    continue

                i = 0
                COMMENTS_WET = list()
                COMMENTS_DRY = list()
                with pd.ExcelWriter(out_file, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer :
                    temp_qc.to_excel(writer, sheet_name='qc_info',index=True, index_label=False)
                    ws = writer.sheets['qc_info']
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=1, column=col).value
                        if cell_value is None:
                            ws.column_dimensions[get_column_letter(col)].width = 30
                        elif cell_value in ['CTRL','DRY','TS','GS','LD']:
                            ws.column_dimensions[get_column_letter(col)].width = 8
                            ws[get_column_letter(col)+'1'].alignment = Alignment(horizontal = 'center', vertical = 'center', wrapText = True)
                        else :
                            ws.column_dimensions[get_column_letter(col)].width = 20
                            ws[get_column_letter(col)+'1'].alignment = Alignment(horizontal = 'center', vertical = 'center', wrapText = True)


                    for row in ws:
                        i += 1
                        sampleID = ws.cell(row = i, column = 1).value
                        for cell in row:
                            col = ws.cell(row = 1, column = cell.column).value
                            if col is None : continue                   # Skip column names
                            cell.alignment = alignment
                            ws[cell.coordinate].border = border         # Draw a ruled line

                            if i % 2 == 0:
                                ws[cell.coordinate].fill = fill         # Color cells every two rows.

                            if col.endswith('_pass') and not ws[cell.coordinate].value :
                                ws[cell.coordinate].font = font         # Change "FALSE" to bold red
                                COMMENTS_DRY.append(sampleID)
                            elif col.endswith('_GRADE') and ws[cell.coordinate].value == 'FAIL':
                                ws[cell.coordinate].font = font         # Change "FALSE" to bold red
                                COMMENTS_WET.append(sampleID)

                    ws0 = writer.sheets['work_sheet']

                    if item['PRJ_TYPE'] == 'eWES':
                        ws0["K26"] = sum(np.isin(list(set(COMMENTS_DRY)), temp_qc.index[temp_qc['CTRL']=='']))
                        ws0["N26"] = sum(np.isin(list(set(COMMENTS_WET)), temp_qc.index[temp_qc['CTRL']=='']))
                    elif item['PRJ_TYPE'] == 'WTS':
                        ws0["K25"] = sum(np.isin(list(set(COMMENTS_DRY)), temp_qc.index[temp_qc['CTRL']=='']))
                        ws0["N25"] = sum(np.isin(list(set(COMMENTS_WET)), temp_qc.index[temp_qc['CTRL']=='']))


class ADD_STAT:

    def __init__(self, args) :
        self.flowcellid = args.flowcellid
        self.directory = args.directory
        self.project_type = args.project_type
        self.outdir = args.outdir

    def fetchStat_WTS(self, summary_dir:Path, sample_id):
        fs_file = os.path.join( summary_dir, sample_id +'.summarized.fusion.tsv')
        sp_file = os.path.join( summary_dir, sample_id +'.summarized.splice.tsv')

        if os.path.isfile(fs_file) :
            fs_data = pd.read_csv(fs_file, sep="\t")
            fs_data = fs_data[['gene1','gene2','chr1','breakpoint_1','chr2','breakpoint_2','max_split_cnt','max_span_cnt']].drop_duplicates().sort_values('gene1').reset_index(drop=True)
        else :
            fs_data = None

        if os.path.isfile(sp_file) :
            sp_data = pd.read_csv(sp_file, sep="\t")
            sp_data = sp_data[ sp_data['FILTER'] == 'PASS' ][['spliceName','discordant_mates','canonical_reads','ratio','tpm_total','tpm_variant']].drop_duplicates()
        else :
            sp_data = None

        return { 'fusion_data': fs_data, 'AS_data': sp_data }

    def fetchStat_eWES(self, summary_dir:Path, sample_id):
        cnv_file = os.path.join( summary_dir, sample_id + '.summarized.cnv.exome.tsv')
        msi_file = os.path.join( summary_dir, sample_id + '.summarized.msi.exome.tsv')
        tmb_file = os.path.join( summary_dir, sample_id + '.summarized.tmb.exome.tsv')
        target_file = os.path.join( summary_dir, sample_id + '.summarized.snv.target.tsv')
        exome_file = os.path.join( summary_dir, sample_id + '.summarized.snv.exome.tsv')

        for f_path in [cnv_file, msi_file, tmb_file, target_file, exome_file] :
            if not os.path.isfile(f_path) :
                return None, None, None, None, None

        cnv_data = pd.read_csv(cnv_file,sep="\t")
        msi_data = pd.read_csv(msi_file,sep="\t")
        tmb_data = pd.read_csv(tmb_file,sep="\t")
        target_data = pd.read_csv(target_file,sep="\t", low_memory=False, dtype=str)
        exome_data = pd.read_csv(exome_file,sep="\t", low_memory=False, dtype=str)

        msi_data = msi_data[['MSI','Result']].drop_duplicates()
        tmb_data = tmb_data[['TMB','TMB_STATUS']].drop_duplicates()
        cnv_data = cnv_data[ cnv_data['FILTER']=='PASS' ][['Gene_name','TYPE','ONCOKB_ONCOGENICITY','gene.mean.CN']].drop_duplicates().sort_values('Gene_name')
        if cnv_data.shape[0] == 0 : cnv_data = None

        target_data = target_data.infer_objects(copy=False).fillna(np.nan).replace([np.nan], [None])
        target_data = target_data[['SYMBOL','HGVSc','HGVSp','AF','Clinvar_CLNSIG','ONCOKB_ONCOGENICITY']].drop_duplicates()
        target_data["HGVSc"] = target_data["HGVSc"].str.split(":", expand=True)[1]
        target_data["HGVSp"] = target_data["HGVSp"].str.split(":", expand=True)[1]
        target_filt = (target_data["Clinvar_CLNSIG"].str.contains("Pathogenic|Likely_pathogenic", case=True, na=False) | target_data["ONCOKB_ONCOGENICITY"].str.contains("oncogenic", case=False, na=False))
        target_data = target_data.loc[target_filt].sort_values('SYMBOL').reset_index(drop=True)

        exome_data = exome_data.infer_objects(copy=False).fillna(np.nan).replace([np.nan], [None])
        exome_data["Clinvar_CLNSIG"] = exome_data["Clinvar_CLNSIG"].str.replace("&", " & ").str.replace("/", " / ")
        exome_data = exome_data.loc[ exome_data["Clinvar_CLNSIG"].str.contains("Pathogenic|Likely_pathogenic", case=True, na=False) ]   # exome checks only ClinVar.
        exome_data = exome_data[ ~exome_data['SYMBOL'].isin(target_data['SYMBOL'])]
        exome_data = exome_data[['SYMBOL']].drop_duplicates().sort_values('SYMBOL').reset_index(drop=True)

        if target_data.shape[0] == 0 : target_data = None
        if exome_data.shape[0] == 0 : exome_data = None

        return {'cnv_data':cnv_data, 'msi_data':msi_data, 'tmb_data':tmb_data, 'target_data':target_data, 'exome_data':exome_data }

    def merge_stat(self, stats_dict, tmp_stats, idx, sample_id):
        if not tmp_stats[idx] is None :
            tmp_stats[idx].insert(0, 'sample_id', sample_id)
            if idx in stats_dict:
                stats_dict[idx] = pd.concat([stats_dict[idx], tmp_stats[idx]], axis=0)
            else :
                stats_dict[idx] = tmp_stats[idx]
        return stats_dict

    def __call__(self):

        df_info = getinfo(self.flowcellid)
        if df_info.shape[0] == 0 : init("flowcell ID not registered.")

        df_info['PRJ_TYPE'] = df_info['PRJ_TYPE'].str.replace('EWES',"eWES")
        if self.project_type == "both" :
            df_info = df_info[ df_info['PRJ_TYPE'].isin(['eWES','WTS']) ]
        else :
            df_info = df_info[ df_info['PRJ_TYPE']==self.project_type ]
        if df_info.shape[0] == 0 : init("No data for the relevant project type.")

        uniq_info = fcDir_table(df_info, self.directory)
        if uniq_info.shape[0] == 0: init("No analysed samples.")

        df_info = pd.merge(df_info, uniq_info, on=['sub_name','PRJ_TYPE'])

        for i, item1 in uniq_info.iterrows() :

            out_file = os.path.join(self.outdir, '.'.join([item1['seqDir'], item1['PRJ_TYPE'],'xlsx']))
            if not os.path.isfile(out_file) : init("Worksheet not yet prepared")

            if item1['PRJ_TYPE'] == 'eWES':
                PC = '_PCE_'
                NC = '_NCE_'
            elif item1['PRJ_TYPE'] == 'WTS':
                PC = '_PCT_'
                NC = '_NCT_'
            else :
                print('PRJ_TYPE error:' + item1['PRJ_TYPE'] )
                continue

            tmp_info = df_info[ (df_info['seqDir'] == item1['seqDir']) & (df_info['PRJ_TYPE'] == item1['PRJ_TYPE']) ]
            tmp_info = tmp_info[ ~tmp_info['SAMPLE_ID'].str.contains(PC, case=False, regex=False) ]
            tmp_info = tmp_info[ ~tmp_info['SAMPLE_ID'].str.contains(NC, case=False, regex=False) ]

            tmp_dict = {}
            for j, item2 in tmp_info.iterrows():

                summary_dir = Path( os.path.join(self.directory, item2['PRJ_TYPE'], item2['seqDir'], item2['SAMPLE_ID'], 'Summary') )
                if not os.path.isfile(os.path.join(summary_dir, item2['SAMPLE_ID'] + '.report.json')): continue
                tmp_stats = getattr(self, f"fetchStat_{item2['PRJ_TYPE']}")(summary_dir, item2['SAMPLE_ID'])
                for idx in tmp_stats.keys() :
                    tmp_dict = self.merge_stat(tmp_dict, tmp_stats, idx, item2['SAMPLE_ID'])

            for idx in tmp_dict.keys() :
                if tmp_dict[idx] is None : continue
                with pd.ExcelWriter(out_file, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer :
                    tmp_dict[idx].to_excel(writer, sheet_name=idx, index=False, index_label=False)

                    # Adjustment of column width
                    ws = writer.sheets[idx]
                    for col in ws.columns:
                        max_length = 0
                        col_letter = col[0].column_letter
                        i = 0
                        for cell in col:
                            i += 1
                            cell.alignment = alignment
                            ws[cell.coordinate].border = border                     # Draw a ruled line
                            if i % 2 == 0: ws[cell.coordinate].fill = fill          # Color cells every two rows.

                            try:
                                max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                            ws.column_dimensions[col_letter].width = max_length + 5


def additional_worksheet(args):

    thfile = os.path.join(os.path.dirname(__file__), 'thresholds.tsv')
    assert os.path.isfile(thfile)
    tbl_th = pd.read_csv(thfile, sep="\t", index_col=None)

    os.makedirs(args.outdir, exist_ok=True)
    runner = ADD_QC(args)
    runner()

    runner = ADD_STAT(args)
    runner()


