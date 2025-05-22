import os
import sys
import pandas as pd
import numpy as np
import openpyxl
import datetime
import warnings
import shutil
from pathlib import Path
from .commom import *

def prompt_choice(prompt, choices):
    while True:
        ans = input(prompt).strip().lower()
        if ans in choices:
            return ans

def create_worksheet(args):

    flowcellid = args.flowcellid
    directory = args.directory
    project_type = args.project_type
    outdir = args.outdir

    df_info = getinfo(flowcellid)
    if df_info.shape[0] == 0 : self.init("flowcell ID not registered.")

    df_info['PRJ_TYPE'] = df_info['PRJ_TYPE'].str.replace('EWES',"eWES")

    if project_type == "both" :
        df_info[ df_info['PRJ_TYPE'].isin(['eWES','WTS']) ]
    else :
        df_info = df_info[ df_info['PRJ_TYPE']==project_type]
    if df_info.shape[0] == 0 : init("No data for the relevant project type.")

    uniq_info = fcDir_table(df_info, directory)
    if uniq_info.shape[0] == 0: init("No analysed samples.")

    df_info = pd.merge(df_info, uniq_info, on=['sub_name','PRJ_TYPE'])

    for i, item in uniq_info.iterrows() :
        if item['PRJ_TYPE'] == 'eWES':
            PC = '_PCE_'
            NC = '_NCE_'
        elif item['PRJ_TYPE'] == 'WTS':
            PC = '_PCT_'
            NC = '_NCT_'
        else :
            print('PRJ_TYPE error:' + item['PRJ_TYPE'] )
            continue

        tempfile = os.path.join(Path(os.path.abspath(__file__)).parent.parent, 'templete', 'templete.'+item['PRJ_TYPE']+'.xlsx')
        if not os.path.isfile(tempfile) :
            print('Template file not found: ' + tempfile)
            continue

        temp_info = df_info[ (df_info['sub_name']==item['sub_name']) & (df_info['PRJ_TYPE']==item['PRJ_TYPE']) ].reset_index(drop=True)
        temp_info = temp_info[['run_id','PRJ_TYPE','seqDir','ANAL_STATUS','SAMPLE_ID','PATIENT_NO','DIAGNOSIS_NAME','GENDER','AGE']]
        temp_info['CTRL'] = np.where(temp_info['SAMPLE_ID'].str.contains(PC),'PC', np.where(temp_info['SAMPLE_ID'].str.contains(NC),'NC',''))

        out_file = os.path.join(outdir, '.'.join([item['seqDir'], item['PRJ_TYPE'],'xlsx']))
        if os.path.isfile(out_file):
            choice = prompt_choice(out_file + " is exists. Do you want to overwrite it? (yes[Y]/no[N]): ", ['yes', 'y', 'no', 'n'])
            if choice in ['no', 'n']:
                print('Suspend operation.')
                continue

        shutil.copy(tempfile, out_file)

        wb = openpyxl.load_workbook(out_file)
        ws = wb['work_sheet']
        ws["B3"] = temp_info.loc[0,'run_id']
        ws["D3"] = temp_info.loc[0,'seqDir']
        ws["L3"] = temp_info.loc[0,'PRJ_TYPE']
        ws["N3"] = temp_info[temp_info['CTRL']==''].shape[0]
        ws["P3"] = "0"
        ws["R3"] = datetime.datetime.now().strftime("%Y/%m/%d %H:%M")
        wb.save(out_file)

        with pd.ExcelWriter(out_file, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer :
            temp_info.to_excel(writer, sheet_name='sample_info',index=False)



